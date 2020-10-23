"""
The MIT License (MIT)
Copyright (c) 2020 cdhsu2001@gmail.com

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of
the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO
THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

Developed by cdhsu2001@gmail.com in 2020
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Side, Border
from openpyxl import Workbook
import re
import time

timerstart = time.time()

XMLfilename = input('Input xml (of STDF) here, its filename (not zipped) should be file.xml: ')
frd = open(XMLfilename, mode='r', encoding='utf-8')
filename = re.search('(.*).xml', XMLfilename)
XLSXfilename = filename.group(1) + 'Map.xlsx'

for i in range(0, 50):  
    line = frd.readline() 
    if line.startswith('SDR:'): 
        SDRsite_cnt = re.search('SITE_CNT=(\d+)', line) 
        SDRsitecnt = int(SDRsite_cnt.group(1)) 
        break

wb = Workbook()
ws1 = wb.active
ws1.title = "AllResult"
xlsxheadrow = 7 
xlsxrow = 8  

OneWaferdict = { 'head_num':[], 'site_num':[], 'hard_bin':[], 'soft_bin':[], 'x_coord':[], 'y_coord':[] }
OneWaferdictFirst = { 'head_num':[], 'site_num':[], 'hard_bin':[], 'soft_bin':[], 'x_coord':[], 'y_coord':[] }
OneWaferdictRetest = { 'head_num':[], 'site_num':[], 'hard_bin':[], 'soft_bin':[], 'x_coord':[], 'y_coord':[] }
OneTDdict = { 'head_num':[], 'site_num':[], 'hard_bin':[], 'soft_bin':[], 'x_coord':[], 'y_coord':[] }
Headerinxlsx = list(OneTDdict.keys()) 
for k in Headerinxlsx:
    for j in range(0, SDRsitecnt):
        OneTDdict[k].append('na')

HBRcont = { 'hbin_num':[], 'hbin_pf':[], 'hbin_nam':[] }
Firstcont = { 'hbin_qty':[], 'hbin_yield':[] }
Retestcont = { 'hbin_qty':[], 'hbin_yield':[] }
PCRcont = { 'part_cnt':0, 'rtst_cnt':0 }

PTRLoLimit = {}
PTRHiLimit = {}
MPRLoLimit = {}
MPRHiLimit = {}
PIRindex = 0    
PRRindex = 0

while ( line != ''):
    line = frd.readline()
    
    if line.startswith('PIR:'): 
        PIRhead = re.search('HEAD_NUM=(\d+)', line)
        OneTDdict['head_num'][PIRindex] = int(PIRhead.group(1))
        PIRsite = re.search('SITE_NUM=(\d+)', line)
        OneTDdict['site_num'][PIRindex] = int(PIRsite.group(1))
        PIRindex += 1
      
    elif line.startswith('PTR:'):     
        PTRresult = re.search('RESULT=(.*) TEST_TXT', line)
        PTRrslt = float( PTRresult.group(1) )
        PIRcount = PIRindex 
        PTRtest_num = re.search('TEST_NUM=(\d+)', line)
        testnumkey = 'test_num '+ PTRtest_num.group(1)
        if (testnumkey in OneTDdict) == False:
            OneTDdict[testnumkey] = []
            for j in range(0, SDRsitecnt):
                OneTDdict[testnumkey].append('na')        
###
        PTRopt_flag = re.search('OPT_FLAG=(\d+)', line)
        if PTRopt_flag:
            PTRoptflag = int( PTRopt_flag.group(1) )
        
            if (testnumkey in PTRLoLimit) == False:       
                PTRLoLimit[testnumkey] = 'na'
                if (((PTRoptflag >> 4) & 0x1) == 0) and (((PTRoptflag >> 6) & 0x1) == 0):
                    PTRlo_limit = re.search('LO_LIMIT=(.*) HI_LIMIT', line) 
                    PTRLoLimit[testnumkey] = float( PTRlo_limit.group(1) )
                                
            if (testnumkey in PTRHiLimit) == False:       
                PTRHiLimit[testnumkey] = 'na'                
                if (((PTRoptflag >> 5) & 0x1) == 0) and (((PTRoptflag >> 7) & 0x1) == 0):
                    PTRhi_limit = re.search('HI_LIMIT=(.*) UNITS', line) 
                    PTRHiLimit[testnumkey] = float( PTRhi_limit.group(1) )
###        
        PTRhead = re.search('HEAD_NUM=(\d+)', line)
        PTRsite = re.search('SITE_NUM=(\d+)', line)
        PTRtestflg = re.search('TEST_FLG=(\d+)', line)
        for j in range(0, PIRcount):
            if (int(PTRhead.group(1)) == OneTDdict['head_num'][j]) and (int(PTRsite.group(1)) == OneTDdict['site_num'][j]):
                if int(PTRtestflg.group(1)) == 0 :
                    OneTDdict[testnumkey][j] = PTRrslt        
        
    elif line.startswith('MPR:'): 
        MPRrslt_cnt = re.search('RSLT_CNT=(\d+)', line)
        MPRrsltcnt = int( MPRrslt_cnt.group(1) )
        if MPRrsltcnt != 1:
            print('MPR RSLT_CNT is not equal to 1, this script will not support!')
            break

        MPRrtn_rslt = re.search('RTN_RSLT=(.*) TEST_TXT', line)
        MPRrtnrslt = eval( MPRrtn_rslt.group(1) ) 
        PIRcount = PIRindex 
        MPRtest_num = re.search('TEST_NUM=(\d+)', line)
        testnumkey = 'test_num '+ MPRtest_num.group(1)
        if (testnumkey in OneTDdict) == False:
            OneTDdict[testnumkey] = []
            for j in range(0, SDRsitecnt):
                OneTDdict[testnumkey].append('na')
###
        MPRopt_flag = re.search('OPT_FLAG=(\d+)', line)
        if MPRopt_flag:
            MPRoptflag = int( MPRopt_flag.group(1) )               
                
            if (testnumkey in MPRLoLimit) == False:       
                MPRLoLimit[testnumkey] = 'na'
                if (((MPRoptflag >> 4) & 0x1) == 0) and (((MPRoptflag >> 6) & 0x1) == 0):
                    MPRlo_limit = re.search('LO_LIMIT=(.*) HI_LIMIT', line)  
                    MPRLoLimit[testnumkey] = float( MPRlo_limit.group(1) )
                                
            if (testnumkey in MPRHiLimit) == False:       
                MPRHiLimit[testnumkey] = 'na'                
                if (((MPRoptflag >> 5) & 0x1) == 0) and (((MPRoptflag >> 7) & 0x1) == 0):
                    MPRhi_limit = re.search('HI_LIMIT=(.*) START_IN', line) 
                    MPRHiLimit[testnumkey] = float( MPRhi_limit.group(1) )
###
        MPRhead = re.search('HEAD_NUM=(\d+)', line)
        MPRsite = re.search('SITE_NUM=(\d+)', line)
        MPRtestflg = re.search('TEST_FLG=(\d+)', line)
        for j in range(0, PIRcount):
            if (int(MPRhead.group(1)) == OneTDdict['head_num'][j]) and (int(MPRsite.group(1)) == OneTDdict['site_num'][j]):
                if int(MPRtestflg.group(1)) == 0 : 
                    OneTDdict[testnumkey][j] = float(MPRrtnrslt[0])       
        
    elif line.startswith('PRR:'): 
        PIRcount = PIRindex 
        PRRhead = re.search('HEAD_NUM=(\d+)', line)
        PRRsite = re.search('SITE_NUM=(\d+)', line)
        PRRhbin = re.search('HARD_BIN=(\d+)', line)
        PRRsbin = re.search('SOFT_BIN=(\d+)', line)
        PRRxcrd = re.search('X_COORD=([-+]?\d+)', line)
        PRRycrd = re.search('Y_COORD=([-+]?\d+)', line)
        
        for j in range(0, PIRcount):
            if (int(PRRhead.group(1)) == OneTDdict['head_num'][j]) and (int(PRRsite.group(1)) == OneTDdict['site_num'][j]):
                OneTDdict['hard_bin'][j] = int(PRRhbin.group(1))
                OneTDdict['soft_bin'][j] = int(PRRsbin.group(1))
                OneTDdict['x_coord'][j] = int(PRRxcrd.group(1))
                OneTDdict['y_coord'][j] = int(PRRycrd.group(1))
                PRRindex += 1
                OneWaferdict['head_num'].append( int(PRRhead.group(1)) )
                OneWaferdict['site_num'].append( int(PRRsite.group(1)) )
                OneWaferdict['hard_bin'].append( int(PRRhbin.group(1)) )
                OneWaferdict['soft_bin'].append( int(PRRsbin.group(1)) )
                OneWaferdict['x_coord'].append( int(PRRxcrd.group(1)) )
                OneWaferdict['y_coord'].append( int(PRRycrd.group(1)) )
            
        if PIRcount == PRRindex:
            OneTDheader = list(OneTDdict.keys())
            if len(Headerinxlsx) < len(OneTDheader):
                Headerinxlsx = OneTDheader 
                ws1.cell( row=1, column=1, value='Hi_Limit' ) ###
                ws1.cell( row=2, column=1, value='Lo_Limit' ) ###
                for y in range(0, len(OneTDheader)):
                    ws1.cell( row=xlsxheadrow, column=(y+1), value=OneTDheader[y] )
                    if OneTDheader[y] in PTRHiLimit:
                        ws1.cell( row=1, column=(y+1), value=PTRHiLimit[OneTDheader[y]] )
                    if OneTDheader[y] in PTRLoLimit:
                        ws1.cell( row=2, column=(y+1), value=PTRLoLimit[OneTDheader[y]] )
                    if OneTDheader[y] in MPRHiLimit:
                        ws1.cell( row=1, column=(y+1), value=MPRHiLimit[OneTDheader[y]] )
                    if OneTDheader[y] in MPRLoLimit:
                        ws1.cell( row=2, column=(y+1), value=MPRLoLimit[OneTDheader[y]] )
                
            for y in range(0, len(OneTDheader)):
                keyname = OneTDheader[y]
                for x in range(0, PIRcount): 
                    ws1.cell( row=(xlsxrow+x), column=(y+1), value=OneTDdict[keyname][x] )  
            xlsxrow = xlsxrow + PIRcount  
            
            for k in OneTDdict.keys():
                for j in range(0, SDRsitecnt):
                    OneTDdict[k][j] = 'na' 
            PIRindex = 0 
            PRRindex = 0 
###
###            
    if line.startswith('HBR:'): 
        HBRhead = re.search('HEAD_NUM=(\d+)', line)
        if int(HBRhead.group(1)) == 255:
            HBRhbin_num = re.search('HBIN_NUM=(\d+)', line)
            HBRcont['hbin_num'].append( int(HBRhbin_num.group(1)) )
            HBRhbin_pf = re.search('HBIN_PF=([PF]+)', line)
            HBRcont['hbin_pf'].append( HBRhbin_pf.group(1) )
            HBRhbin_nam = re.search('HBIN_NAM=(.*)', line)
            if HBRhbin_nam:
                HBRcont['hbin_nam'].append( HBRhbin_nam.group(1) )
            else:
                HBRcont['hbin_nam'].append('na')
###
###                
    if line.startswith('PCR:'): 
        PCRhead = re.search('HEAD_NUM=(\d+)', line)
        if int(PCRhead.group(1)) == 255:
            PCRpart_cnt = re.search('PART_CNT=(\d+)', line)
            PCRcont['part_cnt'] = int(PCRpart_cnt.group(1))
            PCRrtst_cnt = re.search('RTST_CNT=(\d+)', line)
            PCRcont['rtst_cnt'] = int(PCRrtst_cnt.group(1))
###
###            
print('PCR: HEAD_NUM=255', PCRcont)
print('HBR: HEAD_NUM=255', HBRcont)
print('OneWaferdict length', len(OneWaferdict['site_num']) )
print('OneWaferdict Xmax', max(OneWaferdict['x_coord']) , 'Xmin', min(OneWaferdict['x_coord']) )
print('OneWaferdict Ymax', max(OneWaferdict['y_coord']) , 'Ymin', min(OneWaferdict['y_coord']) )
OneWaferXmax = max(OneWaferdict['x_coord'])
OneWaferXmin = min(OneWaferdict['x_coord'])
OneWaferYmax = max(OneWaferdict['y_coord'])
OneWaferYmin = min(OneWaferdict['y_coord'])
###
Hbincolor = []
for k in range(0, 40, 1):
    Hbincolor.append('na')

Hbincolor[0] = '00FFFFFF' 
Hbincolor[1] = '0000FF00' 
Hbincolor[2] = '0099CC00' 
Hbincolor[3] = '00008000' 
Hbincolor[4] = '00003300' 
Hbincolor[6] = '00FF00FF' 
Hbincolor[7] = '0000FFFF' 
Hbincolor[10] = '000000FF' 
Hbincolor[11] = '00FF0000' 
Hbincolor[12] = '00FFFF00' 
Hbincolor[13] = '00000080' 
Hbincolor[15] = '00800000' 
Hbincolor[17] = '00808000' 
Hbincolor[18] = '00800080' 
Hbincolor[19] = '00008080' 
Hbincolor[20] = '009999FF' 
Hbincolor[22] = '00993366' 
Hbincolor[23] = '00FFFFCC' 
Hbincolor[24] = '00CCFFFF' 
Hbincolor[25] = '00660066' 
Hbincolor[26] = '00FF8080' 
Hbincolor[27] = '00C0C0C0' 
Hbincolor[28] = '00808080' 
Hbincolor[29] = '000066CC' 
Hbincolor[30] = '00CCCCFF' 
Hbincolor[31] = '00FFFF99' 
Hbincolor[34] = '0099CCFF' 
###

# only consider one head (HEAD_NUM=1) from STDF
MapPosX = 'L' # can be toward 'R' (right) or 'L' (left)
MapPosY = 'D' # can be toward 'U' (up) or 'D' (down)
colstart = 3
rowstart = 3 
sd = Side(style='medium', color="000000")
blackbd= Border(left=sd, top=sd, right=sd, bottom=sd)


ws2 = wb.create_sheet(title="FirstscreenedMap")
if PCRcont['rtst_cnt'] > 0:
    ws3 = wb.create_sheet(title="RetestedMap")
retest_count = 0
    
if MapPosX == 'L':
    Xgapvalue = OneWaferXmax + colstart
    colindex = colstart
    for i in range(OneWaferXmax, (OneWaferXmin-1), -1):
        ws2.column_dimensions[get_column_letter(colindex)].width = 5.0
        ws2.cell( row=1, column=colindex, value=i ).border = blackbd
        if PCRcont['rtst_cnt'] > 0:
            ws3.column_dimensions[get_column_letter(colindex)].width = 5.0
            ws3.cell( row=1, column=colindex, value=i ).border = blackbd
        colindex += 1 

if MapPosX == 'R':
    Xgapvalue = OneWaferXmin - colstart # to be checked
    colindex = colstart
    for i in range(OneWaferXmin, (OneWaferXmax+1), 1):
        ws2.column_dimensions[get_column_letter(colindex)].width = 5.0
        ws2.cell( row=1, column=colindex, value=i ).border = blackbd
        colindex += 1 
        
if MapPosY == 'D':
    Ygapvalue = OneWaferYmin - rowstart
    rowindex = rowstart
    for i in range(OneWaferYmin, (OneWaferYmax+1), 1):
        ws2.column_dimensions['A'].width = 5.0
        ws2.cell( row=rowindex, column=1, value=i ).border = blackbd
        if PCRcont['rtst_cnt'] > 0:
            ws3.column_dimensions['A'].width = 5.0
            ws3.cell( row=rowindex, column=1, value=i ).border = blackbd
        rowindex += 1 

if MapPosY == 'U':
    Ygapvalue = OneWaferYmax + rowstart # to be checked
    rowindex = rowstart
    for i in range(OneWaferYmax, (OneWaferYmin-1), -1):
        ws2.column_dimensions['A'].width = 5.0
        ws2.cell( row=rowindex, column=1, value=i ).border = blackbd
        rowindex += 1 

for j in range(0, (PCRcont['part_cnt']+PCRcont['rtst_cnt']), 1):
    if (MapPosX == 'L') and (MapPosY == 'D'):
        colnow = Xgapvalue - OneWaferdict['x_coord'][j]
        rownow = OneWaferdict['y_coord'][j] - Ygapvalue
        if j<PCRcont['part_cnt']:
            ws2.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).border = blackbd
            if PCRcont['rtst_cnt'] > 0:
                ws3.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).border = blackbd
        elif j>=PCRcont['part_cnt']:
            ws3.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).border = blackbd
        
    if Hbincolor[ OneWaferdict['hard_bin'][j] ] != 'na':
        colorindex = Hbincolor[ OneWaferdict['hard_bin'][j] ]
        fillcolor = PatternFill(fill_type='solid', start_color=colorindex)
        if j<PCRcont['part_cnt']:
            ws2.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).fill = fillcolor
            if PCRcont['rtst_cnt'] > 0:
                ws3.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).fill = fillcolor
        elif j>=PCRcont['part_cnt']:
            ws3.cell( row=rownow, column=colnow, value=OneWaferdict['hard_bin'][j] ).fill = fillcolor

    if j<PCRcont['part_cnt']:
        for k in OneWaferdict.keys():
            OneWaferdictFirst[k].append( OneWaferdict[k][j] )
            OneWaferdictRetest[k].append( OneWaferdict[k][j] )
    elif j>=PCRcont['part_cnt']:   
        for m in range(0, PCRcont['part_cnt'], 1):
            if (OneWaferdict['x_coord'][j]==OneWaferdictRetest['x_coord'][m]) and (OneWaferdict['y_coord'][j]==OneWaferdictRetest['y_coord'][m]):
                OneWaferdictRetest['hard_bin'][m] = OneWaferdict['hard_bin'][j]
                OneWaferdictRetest['soft_bin'][m] = OneWaferdict['soft_bin'][j]
                retest_count += 1
                break ###
        
if (MapPosX == 'L') and (MapPosY == 'D'):
    bincolnow = Xgapvalue - OneWaferXmin + 1
    binrownow = OneWaferYmin - Ygapvalue
        
    for i in range(0, len(HBRcont['hbin_num']), 1):
        Firstcont['hbin_qty'].append( OneWaferdictFirst['hard_bin'].count( HBRcont['hbin_num'][i] ) )        
        Firstcont['hbin_yield'].append( (OneWaferdictFirst['hard_bin'].count(HBRcont['hbin_num'][i])) / PCRcont['part_cnt'] ) 
        colorindex = Hbincolor[ HBRcont['hbin_num'][i] ]
        fillcolor = PatternFill(fill_type='solid', start_color=colorindex)
        ws2.cell( row=binrownow, column=bincolnow, value=HBRcont['hbin_num'][i] ).fill = fillcolor
        ws2.cell( row=binrownow, column=bincolnow+1, value=HBRcont['hbin_nam'][i] )
        ws2.cell( row=binrownow, column=bincolnow+2, value=Firstcont['hbin_qty'][i] )
        ws2.cell( row=binrownow, column=bincolnow+3, value=Firstcont['hbin_yield'][i] )
        if PCRcont['rtst_cnt'] > 0:
            Retestcont['hbin_qty'].append( OneWaferdictRetest['hard_bin'].count( HBRcont['hbin_num'][i] ) )        
            Retestcont['hbin_yield'].append( (OneWaferdictRetest['hard_bin'].count(HBRcont['hbin_num'][i])) / PCRcont['part_cnt'] )
            ws3.cell( row=binrownow, column=bincolnow, value=HBRcont['hbin_num'][i] ).fill = fillcolor
            ws3.cell( row=binrownow, column=bincolnow+1, value=HBRcont['hbin_nam'][i] )            
            ws3.cell( row=binrownow, column=bincolnow+2, value=Retestcont['hbin_qty'][i] )
            ws3.cell( row=binrownow, column=bincolnow+3, value=Retestcont['hbin_yield'][i] )
        binrownow += 1 
       
        
print('retest_count=', retest_count)
frd.close()
wb.save(XLSXfilename)
timerend = time.time()
print("It spent {}sec".format(timerend-timerstart) )
