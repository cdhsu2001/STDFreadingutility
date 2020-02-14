"""
The MIT License (MIT)
Copyright (c) 2020 cdhsu2001@gmail.com

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of
the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO
THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

Developed by cdhsu2001@gmail.com in 2020
"""

from openpyxl import Workbook
import re
import time

timerstart = time.time()

XMLfilename = input('Input xml (of STDF) here, its filename (not zipped) should be file.xml: ')
frd = open(XMLfilename, mode='r', encoding='utf-8')
filename = re.search('(.*).xml', XMLfilename)
XLSXfilename = filename.group(1) + 'All.xlsx'

for i in range(0, 50):
    line = frd.readline() 
    if line.startswith('SDR:'): 
        SDRsite_cnt = re.search('SITE_CNT=(\d+)', line) 
        SDRsitecnt = int(SDRsite_cnt.group(1)) 
        break

wb = Workbook()
ws1 = wb.active
ws1.title = "AllResult"
xlsxheadrow = 7 
xlsxrow = 8

OneTDdict = { 'head_num':[], 'site_num':[], 'hard_bin':[], 'soft_bin':[], 'x_coord':[], 'y_coord':[] }
Headerinxlsx = list(OneTDdict.keys()) 
for k in Headerinxlsx:
    for j in range(0, SDRsitecnt):
        OneTDdict[k].append('na')

PTRLoLimit = {}
PTRHiLimit = {}
MPRLoLimit = {}
MPRHiLimit = {}
PIRindex = 0    
PRRindex = 0

while ( line != ''):
    line = frd.readline()
    
    if line.startswith('PIR:'): 
        PIRhead = re.search('HEAD_NUM=(\d+)', line)
        OneTDdict['head_num'][PIRindex] = int(PIRhead.group(1))
        PIRsite = re.search('SITE_NUM=(\d+)', line)
        OneTDdict['site_num'][PIRindex] = int(PIRsite.group(1))
        PIRindex += 1
      
    elif line.startswith('PTR:'):   
        PTRresult = re.search('RESULT=(.*) TEST_TXT', line)
        PTRrslt = float( PTRresult.group(1) )
        
        PIRcount = PIRindex 
        PTRtest_num = re.search('TEST_NUM=(\d+)', line)
        testnumkey = 'test_num '+ PTRtest_num.group(1)
        if (testnumkey in OneTDdict) == False:
            OneTDdict[testnumkey] = []
            for j in range(0, SDRsitecnt):
                OneTDdict[testnumkey].append('na')        
###
        PTRopt_flag = re.search('OPT_FLAG=(\d+)', line)        
        if PTRopt_flag:
            PTRoptflag = int( PTRopt_flag.group(1) )
        
            if (testnumkey in PTRLoLimit) == False:       
                PTRLoLimit[testnumkey] = 'na'
                if (((PTRoptflag >> 4) & 0x1) == 0) and (((PTRoptflag >> 6) & 0x1) == 0):
                    PTRlo_limit = re.search('LO_LIMIT=(.*) HI_LIMIT', line)  
                    PTRLoLimit[testnumkey] = float( PTRlo_limit.group(1) )
                                
            if (testnumkey in PTRHiLimit) == False:       
                PTRHiLimit[testnumkey] = 'na'                
                if (((PTRoptflag >> 5) & 0x1) == 0) and (((PTRoptflag >> 7) & 0x1) == 0):
                    PTRhi_limit = re.search('HI_LIMIT=(.*) UNITS', line) 
                    PTRHiLimit[testnumkey] = float( PTRhi_limit.group(1) )
###        
        PTRhead = re.search('HEAD_NUM=(\d+)', line)
        PTRsite = re.search('SITE_NUM=(\d+)', line)
        PTRtestflg = re.search('TEST_FLG=(\d+)', line)
        for j in range(0, PIRcount):
            if (int(PTRhead.group(1)) == OneTDdict['head_num'][j]) and (int(PTRsite.group(1)) == OneTDdict['site_num'][j]):
                if int(PTRtestflg.group(1)) == 0 : 
                    OneTDdict[testnumkey][j] = PTRrslt        
        
    elif line.startswith('MPR:'): 
        MPRrslt_cnt = re.search('RSLT_CNT=(\d+)', line)
        MPRrsltcnt = int( MPRrslt_cnt.group(1) )
        if MPRrsltcnt != 1:
            print('MPR RSLT_CNT is not equal to 1, this script will not support!')
            raise StopIteration

        MPRrtn_rslt = re.search('RTN_RSLT=(.*) TEST_TXT', line)
        MPRrtnrslt = eval( MPRrtn_rslt.group(1) ) 
        
        PIRcount = PIRindex 
        MPRtest_num = re.search('TEST_NUM=(\d+)', line)
        testnumkey = 'test_num '+ MPRtest_num.group(1)
        if (testnumkey in OneTDdict) == False:
            OneTDdict[testnumkey] = []
            for j in range(0, SDRsitecnt):
                OneTDdict[testnumkey].append('na')
###
        MPRopt_flag = re.search('OPT_FLAG=(\d+)', line)
        if MPRopt_flag:
            MPRoptflag = int( MPRopt_flag.group(1) )               
                
            if (testnumkey in MPRLoLimit) == False:       
                MPRLoLimit[testnumkey] = 'na'
                if (((MPRoptflag >> 4) & 0x1) == 0) and (((MPRoptflag >> 6) & 0x1) == 0):
                    MPRlo_limit = re.search('LO_LIMIT=(.*) HI_LIMIT', line) 
                    MPRLoLimit[testnumkey] = float( MPRlo_limit.group(1) )
                                
            if (testnumkey in MPRHiLimit) == False:       
                MPRHiLimit[testnumkey] = 'na'                
                if (((MPRoptflag >> 5) & 0x1) == 0) and (((MPRoptflag >> 7) & 0x1) == 0):
                    MPRhi_limit = re.search('HI_LIMIT=(.*) START_IN', line) 
                    MPRHiLimit[testnumkey] = float( MPRhi_limit.group(1) )
###
        MPRhead = re.search('HEAD_NUM=(\d+)', line)
        MPRsite = re.search('SITE_NUM=(\d+)', line)
        MPRtestflg = re.search('TEST_FLG=(\d+)', line)
        for j in range(0, PIRcount):
            if (int(MPRhead.group(1)) == OneTDdict['head_num'][j]) and (int(MPRsite.group(1)) == OneTDdict['site_num'][j]):
                if int(MPRtestflg.group(1)) == 0 :
                    OneTDdict[testnumkey][j] = float(MPRrtnrslt[0])    
        
    elif line.startswith('PRR:'): 
        PIRcount = PIRindex 
        PRRhead = re.search('HEAD_NUM=(\d+)', line)
        PRRsite = re.search('SITE_NUM=(\d+)', line)
        PRRhbin = re.search('HARD_BIN=(\d+)', line)
        PRRsbin = re.search('SOFT_BIN=(\d+)', line)
        PRRxcrd = re.search('X_COORD=([-+]?\d+)', line)
        PRRycrd = re.search('Y_COORD=([-+]?\d+)', line)
        
        for j in range(0, PIRcount):
            if (int(PRRhead.group(1)) == OneTDdict['head_num'][j]) and (int(PRRsite.group(1)) == OneTDdict['site_num'][j]):
                OneTDdict['hard_bin'][j] = int(PRRhbin.group(1))
                OneTDdict['soft_bin'][j] = int(PRRsbin.group(1))
                OneTDdict['x_coord'][j] = int(PRRxcrd.group(1))
                OneTDdict['y_coord'][j] = int(PRRycrd.group(1))
                PRRindex += 1
            
        if PIRcount == PRRindex:
            OneTDheader = list(OneTDdict.keys())
            if len(Headerinxlsx) < len(OneTDheader):
                Headerinxlsx = OneTDheader 
                ws1.cell( row=1, column=1, value='Hi_Limit' ) 
                ws1.cell( row=2, column=1, value='Lo_Limit' ) 
                for y in range(0, len(OneTDheader)):
                    ws1.cell( row=xlsxheadrow, column=(y+1), value=OneTDheader[y] )
                    if OneTDheader[y] in PTRHiLimit:
                        ws1.cell( row=1, column=(y+1), value=PTRHiLimit[OneTDheader[y]] )
                    if OneTDheader[y] in PTRLoLimit:
                        ws1.cell( row=2, column=(y+1), value=PTRLoLimit[OneTDheader[y]] )
                    if OneTDheader[y] in MPRHiLimit:
                        ws1.cell( row=1, column=(y+1), value=MPRHiLimit[OneTDheader[y]] )
                    if OneTDheader[y] in MPRLoLimit:
                        ws1.cell( row=2, column=(y+1), value=MPRLoLimit[OneTDheader[y]] )
                
            for y in range(0, len(OneTDheader)):
                keyname = OneTDheader[y]
                for x in range(0, PIRcount): 
                    ws1.cell( row=(xlsxrow+x), column=(y+1), value=OneTDdict[keyname][x] )  
            xlsxrow = xlsxrow + PIRcount  
            
            for k in OneTDdict.keys():
                for j in range(0, SDRsitecnt):
                    OneTDdict[k][j] = 'na' 
            PIRindex = 0 
            PRRindex = 0 


frd.close()
wb.save(XLSXfilename)
timerend = time.time()
print("It spent {}sec".format(timerend-timerstart) )
